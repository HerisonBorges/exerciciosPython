import openpyxl


listaNomes = []
wb = openpyxl.load_workbook('dadosFuncionarios.xlsx')

print(wb.sheetnames) 

planilha = wb['Planilha1']


conjunto = tuple(planilha['B2':'D79'])

for linha in conjunto:
    nome = linha[0].value
    salario = linha[2].value

    salarioPlanilha = float(salario)

    if salarioPlanilha > 5000:

        listaNomes.append((nome, salarioPlanilha))
print(listaNomes)


novaPlanilha = wb.create_sheet(title='Planilha2')

novaPlanilha.append(['Nome', 'Salario'])

for nome, salario in listaNomes:

    novaPlanilha.append([nome, salario])

wb.save('dadosFuncionarios.xlsx')






'''book = openpyxl.Workbook()
novaPlanilha = book.active
novaPlanilha.title = 'Funcionarios5000'


novaPlanilha.append(['Nome', 'Salario'])

for nome, salario in listaNomes:

    novaPlanilha.append([nome, salario])

book.save('Funcionarios5000.xlsx')'''