import openpyxl

book = openpyxl.Workbook()

planilha = book.active
planilha.title = "Planilha1"

planilha.append(['Produto', 'Quantidade', 'Preço'])

produtos = [["ARROZ", 5, 5.00],
            ['BATATA', 5, 7.00],
            ['BANANA', 8, 9.00]            
]



for produto in produtos:
    planilha.append(produto)

book.save('ProdutosExercicio.xlsx')

#______________________________________________________________

book = openpyxl.load_workbook('ProdutosExercicio.xlsx')

planilha = book.active

planilha.insert_cols(4)

planilha.cell(row=1, column=4, value='Soma')



for row in range(2, planilha.max_row+1):
    produtoNome = planilha.cell(row=row, column=1).value
    preco = planilha.cell(row=row, column=3).value

    quantidadeCompra = int(input(f"Digite a quandidade de {produtoNome} que vc deseja comprar: "))

    soma = quantidadeCompra * preco

    planilha.cell(row=row, column=4, value=soma)

book.save('ProdutosExercicio.xlsx')

#________________________________________________________