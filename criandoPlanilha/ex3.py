import openpyxl

book = openpyxl.Workbook()

print(book.sheetnames)
      
book.create_sheet('InserirNum')

planilha = book['InserirNum']

numero = int(input('Digite um numero: '))

mult = numero

for row in range(1,11):
    planilha.cell(row=row, column=1).value=row
    planilha.cell(row=row, column=2, value=row*mult)


book.save("InserirNum.xlsx")