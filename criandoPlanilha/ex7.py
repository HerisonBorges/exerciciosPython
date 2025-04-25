import openpyxl

book = openpyxl.Workbook()

book.create_sheet('Teste')

planilha = book['Teste']

planilha.append(["Produto", "Quantidade", "Preço"])

book.save('Teste.xlsx')

#___________________________________PlanihaCriada

book = openpyxl.load_workbook('Teste.xlsx')

print(book.sheetnames)

planilha = book['Teste']

planilha.append(["Arroz", "10", "20.00"])
planilha.append(["Feijão", "12", "30.00"])
planilha.append(["Batata", "30", "15.00"])
planilha.append(["Farinha", "10", "40.00"])

book.save('Teste.xlsx')

#__________________________________________LendoEAdicionando


book= openpyxl.load_workbook('Teste.xlsx')

planilha = book['Teste']

conjunto = tuple(planilha["A2":"D5"])

for linha in conjunto:
    produto = linha[0].value
    quantidade = linha[1].value
    preço = linha[2].value

    print(produto,quantidade,preço)

#_________________________________________lendoosdados

book = openpyxl.load_workbook('Teste.xlsx')

planilha = book["Teste"]

conjunto = tuple(planilha["A2":"D5"])

for linha in conjunto:
    produto = linha[0].value
    quantidade = linha[1].value
    preço = linha[2].value

    preçoPlanilha = float(preço)

    if preçoPlanilha > 20:

        print(produto)

#__________________________________________lendoProdutosMaiorq20

bookMES = openpyxl.Workbook()

meses = ['Janeiro','Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']


for mes in meses:
    planilha = bookMES.create_sheet(title="Mes")


    planilha.append(['Dia', 'Descrição', 'Valor'])

bookMES.save('Meses.xlsx')

#_______________________________________________________________
book = openpyxl.load_workbook('Teste.xlsx')
lista = []
planilha = book["Teste"]

conjunto = tuple(planilha["A2":"D5"])

for linha in conjunto:
    produto = linha[0].value
    quantidade = linha[1].value
    preço = linha[2].value
lista.append((produto,quantidade,preço))



bookCopia = openpyxl.Workbook()

novaAba = bookCopia.active
novaAba=bookCopia.create_sheet(title="BackUP")

novaAba.append(["Produto", "Quantidade", "Preço"])

for produto, quantidade, preço in lista:
    novaAba.append([produto, quantidade, preço])

bookCopia.save("Copia_dados.xlsx")






