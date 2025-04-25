import openpyxl

book = openpyxl.Workbook()

print(book.sheetnames)

book.create_sheet('Copia')
book.create_sheet('Copia2')

planilha = book['Copia']

for row in range(1,11):
    planilha.cell(row=row, column=1).value=row


planilha2 = book['Copia2']

for row in range(1,11):
    planilha2.cell(row=row, column=1).value = planilha.cell(row=row, column=1).value=row

book.save('Copia.xlsx')