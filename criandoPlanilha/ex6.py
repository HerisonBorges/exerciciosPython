import openpyxl

book = openpyxl.Workbook()
book.create_sheet('clientes')

planilha = book['clientes']

planilha.append(['Nome', 'Idade', 'Cidade'])
planilha.append(['Herison', '32', 'uberlandia'])
planilha.append(['Thoor', '24', 'uberlandia'])
planilha.append(['Adriana', '51', 'uberlandia'])



book.save('clientes.xlsx')


wb = openpyxl.load_workbook('clientes.xlsx')

planilhawb = wb['clientes']

conjunto =  tuple(planilhawb['A2':'C4'])

for linha in conjunto:
    nome = linha[0].value
    idade = linha[1].value

    print(nome, idade)


book = openpyxl.load_workbook('clientes.xlsx')

planilha = book['clientes']

planilha['D1'] = 'E-mail'
planilha['D2'] = 'abcd@gmail.com'
planilha['D3'] = 'abbb@gmail.com'
planilha['D4'] = 'accc@gmail.com'


book.save('clientesAtualizados.xlsx')

listaNomes = []
book = openpyxl.load_workbook('clientes.xlsx')

planilha = book['clientes']

conjunto = tuple(planilha['A2':'C4'])

for linha in conjunto:
    nome = linha[0].value
    idade = linha[1].value

    idadeCliente = int(idade)

    if idadeCliente < 50:

        print((nome, idadeCliente))
        listaNomes.append((nome, idadeCliente))
print(listaNomes)



novoBOOK = openpyxl.Workbook()
novaPlanilha = novoBOOK.active
novaPlanilha.title = 'ClientesMenos50'

novaPlanilha.append(['NOME', 'IDADE'])

for nome, idade in listaNomes:
    novaPlanilha.append([nome, idade])

novoBOOK.save('ClientesMenos50.xlsx')