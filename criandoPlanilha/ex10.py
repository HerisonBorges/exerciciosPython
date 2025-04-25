import openpyxl

book = openpyxl.Workbook()

planilha =  book.active

planilha.title = 'Produtos'

planilha.append(["PRODUTO", "QUANTIDADE", "PREÇO"])

produtos = [["Arroz", 5, 5.00],["Batata", 6, 7.00],["Feijao", 4, 9.00],["Laranja", 7, 8.00]]

for produto in produtos:
    planilha.append(produto)

book.save('ExercicioGPT.xlsx')

#_____________________________________________________________

book = openpyxl.load_workbook('ExercicioGPT.xlsx')

planilha = book.active

planilha.insert_cols(4)

planilha.cell(row=1, column=4, value="SOMA")

book.save('ExercicioGPT.xlsx')


#___________________________________________________________

book = openpyxl.load_workbook('ExercicioGPT.xlsx')

planilha = book.active

buscarProduto = input('Digite o nome do produto: ').lower()

produto_encontrado = False

for cel in planilha['A']:
    if cel.value.lower() == buscarProduto:
        produto_encontrado = True
        preco = planilha.cell(row=cel.row, column=3).value
        print(f"O preço do produto {buscarProduto} é de {preco}:")

        break
if not produto_encontrado:
    print("esse produto n existe")


while True:
    buscarProduto = input('Digite o nome do produto ou "sair" para encerrar: ').lower()
    
    # Condição de saída
    if buscarProduto == "sair":
        break

    produto_encontrado = False
    quantidade = int(input("Digite a quantidade: "))

    # Loop através da coluna A para encontrar o produto
    for cel in planilha['A']:
        if cel.value.lower() == buscarProduto:
            produto_encontrado = True
            preco = planilha.cell(row=cel.row, column=3).value
            estoque = planilha.cell(row=cel.row, column=2).value

            if quantidade <= estoque:
                total = quantidade * preco
                print(f"O total de compra de quantidade {quantidade} do produto {buscarProduto}, é de {total:.2f}")
            else: 
                print(f"Desculpe, temos apenas {estoque} em quantidade do produto {buscarProduto}.")
            break

    if not produto_encontrado:
        print("Esse produto não existe.")