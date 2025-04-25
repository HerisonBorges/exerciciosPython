import openpyxl

book = openpyxl.Workbook()
planilha = book.active
planilha.title = 'Produtos'


planilha.append(['Produto', 'Quantidade', 'Preço'])

produtos = [
    ['Arroz', 10, 20.00],
    ['Feijão', 12, 30.00],
    ['Batata', 30, 15.00],
    ['Farinha', 10, 40.00]
]

for produto in produtos:
    planilha.append(produto)


book.save('produtos.xlsx')

#_________________________________________________________

book = openpyxl.load_workbook('produtos.xlsx')

planilha = book['Produtos']

conjunto = tuple(planilha['A2':'C5'])

for cel in conjunto:
    produto = cel[0].value
    quantidade = cel[1].value
    preco = cel[2].value

    print(produto,quantidade,preco)

#_________________________________________________________

book = openpyxl.load_workbook('produtos.xlsx')

planilha = book['Produtos']

totalPreco = sum(cel.value for cel in planilha['C'][1:])

planilha.append(["Total", "", totalPreco])

book.save('produtos.xlsx')

#__________________________________________________________


book = openpyxl.load_workbook('produtos.xlsx')
lista = []
planilha = book['Produtos']

for linha in range(2, planilha.max_row):
    valor = planilha.cell(row=linha, column=1).value
    valor2 = planilha.cell(row=linha, column=3).value
    print(valor)
    print(valor2)

    lista.append((valor, valor2))

print(lista)