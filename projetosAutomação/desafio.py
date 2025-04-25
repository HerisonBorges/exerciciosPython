from openpyxl import load_workbook

def criarAba(bairro, arquivoBairros):
    if bairro not in arquivoBairros.sheetnames:
        # Criando nova aba
        novaAba = arquivoBairros.create_sheet(bairro)
        print(f"Aba '{novaAba}' criada.")
    
    # Garantindo que o cabeçalho sempre seja adicionado ou verificado
    aba = arquivoBairros[bairro]
    
    if aba["A1"].value != "Data de Nascimento":  # Verifica se o cabeçalho está presente
        aba["A1"].value = "Data de Nascimento"
        aba["B1"].value = "Pessoa"
        aba["C1"].value = "Bairro"
        print(f"Cabeçalho adicionado à aba '{bairro}'.")
    else:
        print(f"Cabeçalho já existente na aba '{bairro}'.")


def transferirInformações(abaOrigem, abaDestino, linhaOrigem):
    linhaDestino = abaDestino.max_row + 1
    for coluna in range(1, 4):  # Transferir colunas 1, 2 e 3
        valorOrigem = abaOrigem.cell(row=linhaOrigem, column=coluna).value
        abaDestino.cell(row=linhaDestino, column=coluna).value = valorOrigem

    


# Carregar o arquivo Excel existente
arquivoBairros = load_workbook("Bairros1.xlsx")

# Verificando as abas atuais
print(arquivoBairros.sheetnames)

# Acessando sempre a primeira aba (Base de Dados)
abaBaseDeDados = arquivoBairros["Base de Dados"]

# Acessando o tamanho da planilha (número de linhas)
ultimaLinha = abaBaseDeDados.max_row
print(f"Última linha na base de dados: {ultimaLinha}")

# Percorrendo todas as linhas do 2 até a última linha (excluindo o cabeçalho)
for linha in range(2, ultimaLinha + 1):
    bairro = abaBaseDeDados.cell(row=linha, column=3).value
    
    # Se a célula estiver vazia, ignorar essa linha
    if not bairro:
        print(f"Linha {linha} ignorada (sem bairro).")
        continue

    # Criar uma aba para o bairro, se ainda não existir, e adicionar/verificar o cabeçalho
    criarAba(bairro, arquivoBairros)

    # Transferir as informações para a aba correspondente ao bairro
    abaDestino = arquivoBairros[bairro]
    transferirInformações(abaBaseDeDados, abaDestino, linha)

# Salvar o arquivo Excel com as alterações
arquivoBairros.save("Bairros1.xlsx")
print("Arquivo salvo como 'Bairros1.xlsx'.")
