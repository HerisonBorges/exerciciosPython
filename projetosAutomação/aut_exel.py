from openpyxl import load_workbook

arquivo = load_workbook("alunos.xlsx")

#ver as abas
print(arquivo.sheetnames)

# Selecionar uma aba especifica

aba_alunos = arquivo["Planilha1"]
print(aba_alunos)

#selecionando celulas
valor_A1 = aba_alunos["A1"].value
# ou 
valor_B1 = aba_alunos.cell(row=1, column=2).value
print(valor_B1)

#modificando um valor na planilha

aba_alunos.cell(row=1, column=2).value = "Prova 1"

arquivo.save("Alunos.xlsx")

#descobrindo a ultima linha

print(aba_alunos.max_row)
print(len(aba_alunos["A"]))